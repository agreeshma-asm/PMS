import io
import openpyxl
from copy import copy
from datetime import datetime

TEMPLATE_PATH = r"C:\Users\A30061\Downloads\RC As on 06-04-2026.xlsx"

def generate_traveler_excel(card: dict) -> bytes:
    """Generate a Route Card Excel file based on the standard template."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Header Data Mapping
    # Row 3
    ws.cell(row=3, column=14, value=card.get("workOrderNumber", "")) # WO No
    
    # Row 4
    ws.cell(row=4, column=5, value=card.get("jobName", ""))          # Station / Job Name
    ws.cell(row=4, column=14, value=f"{card.get('batchQuantity', '')} NOS") # Mfg Qty
    
    # Row 5
    ws.cell(row=5, column=5, value=card.get("partRevision", "A"))    # Build / Rev
    ws.cell(row=5, column=14, value=card.get("createdAt", "")[:10])  # Released Date
    
    # Row 6
    ws.cell(row=6, column=5, value=card.get("progNo", ""))           # Prog No
    ws.cell(row=6, column=14, value=card.get("createdBy", ""))       # Released by
    
    # Row 7
    ws.cell(row=7, column=5, value=card.get("projType", ""))         # Project Type
    ws.cell(row=7, column=14, value=card.get("rmGrade", ""))         # RM Grade
    
    # Row 8
    ws.cell(row=8, column=5, value=card.get("partNumber", ""))       # Part No
    ws.cell(row=8, column=8, value=card.get("workCenter", ""))       # Mc Category
    ws.cell(row=8, column=14, value=card.get("rmSize", ""))          # RM Size

    # Clear old steps starting from row 10 and write new ones
    # In the template, steps start at row 10 and are spaced out.
    # To keep it simple, we'll write sequentially starting from row 10.
    start_row = 10
    steps = card.get("steps", [])
    
    for idx, step in enumerate(steps):
        row_idx = start_row + (idx * 2) # Adding spacing between steps like the template
        
        ws.cell(row=row_idx, column=3, value=step.get("stepNumber", ""))
        ws.cell(row=row_idx, column=4, value=step.get("operationName", ""))
        
        # Date
        sign_off_at = step.get("signedOffAt")
        if sign_off_at:
            ws.cell(row=row_idx, column=5, value=sign_off_at[:10])
            
        ws.cell(row=row_idx, column=8, value=step.get("signedOffBy", ""))
        
        # Accepted Qty
        ws.cell(row=row_idx, column=12, value=step.get("completionQty", ""))
        
        # Remarks or status
        status = step.get("status", "")
        if status == "Failed":
            ws.cell(row=row_idx, column=13, value="FAILED")
            ws.cell(row=row_idx, column=15, value=step.get("deviationReason", ""))
        else:
            ws.cell(row=row_idx, column=15, value=status)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
