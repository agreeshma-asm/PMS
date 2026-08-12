"""
Production Management System — PMS Excel Integration
Reads all sheets from PMS.xlsx, computes composite risk scores,
and detects date mismatch alerts across the 7-step process chain.
"""

import openpyxl
import os
from datetime import datetime, timezone, timedelta
import threading
import time

PMS_FILE_PATH = r"C:\Users\A30061\Downloads\PMS.xlsx"

# ─── Standard 7-Step Process Pipeline ──────────────────────────────────────────

STANDARD_PROCESSES = [
    {"stepNumber": 10, "processKey": "iqc",        "operationName": "IQC (Incoming Quality Control)", "workCenter": "QC-IQC",   "instructions": "Verify incoming raw material quality, certifications, and dimensions.", "requiredSop": "SOP-IQC-01"},
    {"stepNumber": 20, "processKey": "rm_cutting",  "operationName": "RM Cutting",                    "workCenter": "CUT-01",   "instructions": "Cut raw material to specified dimensions per drawing.",               "requiredSop": "SOP-CUT-01"},
    {"stepNumber": 30, "processKey": "machining",   "operationName": "Machining",                     "workCenter": "CNC/VMC",  "instructions": "Machine part per CAM program and drawing specifications.",             "requiredSop": "SOP-MC-01"},
    {"stepNumber": 40, "processKey": "deburring",   "operationName": "Deburring",                     "workCenter": "DEB-01",   "instructions": "Remove all burrs and sharp edges. Clean part.",                       "requiredSop": "SOP-DEB-01"},
    {"stepNumber": 50, "processKey": "laser_marking","operationName": "Laser Marking",                "workCenter": "LASER-01", "instructions": "Laser engrave part number, serial number, and traceability marks.",   "requiredSop": "SOP-LAS-01"},
    {"stepNumber": 60, "processKey": "special_process","operationName": "Special Process",             "workCenter": "SP-01",    "instructions": "Heat Treatment / Anodising / Surface Treatment as applicable.",       "requiredSop": "SOP-SP-01"},
    {"stepNumber": 70, "processKey": "qc",          "operationName": "QC (Final Quality Control)",    "workCenter": "QC-OQC",   "instructions": "Final quality inspection — verify all critical dimensions and specs.", "requiredSop": "SOP-QC-01"},
]

# Mapping from PMS Excel columns to process keys
_PROCESS_DATE_COLS = {
    "rm_cutting":      {"date_col": "Cutt Comp_Date",    "qty_col": "Cutt Comp_Qty"},
    "machining":       {"date_col": "Mc Comp_Date",      "qty_col": "Mc Comp_Qty"},
    "deburring":       {"date_col": "Deb Comp_Date",     "qty_col": "Debu Comp_Qty"},
    "laser_marking":   {"date_col": "L-Engr Comp_Date",  "qty_col": "L-Engr Comp_Qty"},
    "qc":              {"date_col": "QC Comp_ Date",     "qty_col": "QC Comp_Qty"},
    "special_process": {"date_col": "SP_Sent Date",      "qty_col": "SP_Sent_ Qty"},
    "iqc":             {"date_col": "IQC Off_ Date",     "qty_col": "IQC Off_Qty"},
}

# ─── Risk Calculation ─────────────────────────────────────────────────────────

def _parse_date(val):
    """Parse a date value from Excel (can be datetime or string)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=None)
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None


def _complexity_score(complexity_str):
    """Convert PMS Complexity string to a numeric score."""
    if not complexity_str:
        return 1  # Default to LOW
    c = str(complexity_str).strip().upper()
    if c in ("HIGH", "HIGH-P", "HIGH-T"):
        return 3
    elif c == "MEDIUM":
        return 2
    else:
        return 1


def _date_proximity_score(target_date):
    """Score based on how close/past the target date is."""
    if not target_date:
        return 1  # No date = LOW
    now = datetime.now()
    days_remaining = (target_date - now).days
    if days_remaining <= 0:
        return 3  # Overdue or due today = HIGH
    elif days_remaining <= 7:
        return 2  # Within 7 days = MEDIUM
    else:
        return 1  # More than 7 days = LOW


def calculate_risk_level(target_date, complexity_str):
    """
    Composite risk = max(Complexity Score, Date Proximity Score)
    Returns: ("HIGH", 3) | ("MEDIUM", 2) | ("LOW", 1)
    """
    c_score = _complexity_score(complexity_str)
    d_score = _date_proximity_score(target_date)
    final_score = max(c_score, d_score)
    level_map = {3: "HIGH", 2: "MEDIUM", 1: "LOW"}
    return level_map[final_score], final_score


# ─── Date Mismatch Alert Detection ────────────────────────────────────────────

def detect_date_mismatches(record):
    """
    Detect date mismatches for a single work order record.
    Returns a list of alert dicts.
    """
    alerts = []
    now = datetime.now()
    target_date = record.get("targetDate")
    processes = record.get("processes", {})

    # 1. Overdue Process — completion date > target date
    if target_date:
        for proc_key, proc_info in processes.items():
            comp_date = proc_info.get("completionDate")
            if comp_date and target_date and comp_date > target_date:
                alerts.append({
                    "type": "OVERDUE_PROCESS",
                    "severity": "HIGH",
                    "process": proc_key,
                    "message": f"{proc_info.get('operationName', proc_key)} completed {(comp_date - target_date).days} days after target date",
                    "targetDate": target_date.isoformat() if target_date else None,
                    "actualDate": comp_date.isoformat() if comp_date else None,
                })

    # 2. Sequence Violation — downstream completed before upstream
    ordered_keys = ["rm_cutting", "machining", "deburring", "laser_marking", "special_process", "qc", "iqc"]
    for i in range(len(ordered_keys) - 1):
        up_key = ordered_keys[i]
        down_key = ordered_keys[i + 1]
        up_date = processes.get(up_key, {}).get("completionDate")
        down_date = processes.get(down_key, {}).get("completionDate")
        if up_date and down_date and down_date < up_date:
            alerts.append({
                "type": "SEQUENCE_VIOLATION",
                "severity": "MEDIUM",
                "process": down_key,
                "message": f"{processes.get(down_key, {}).get('operationName', down_key)} completed before {processes.get(up_key, {}).get('operationName', up_key)}",
                "upstreamDate": up_date.isoformat(),
                "downstreamDate": down_date.isoformat(),
            })

    # 3. Stale WIP — no completion for > 3 days past target
    if target_date and (now - target_date).days > 3:
        status = record.get("status", "")
        if status and status not in ("Closed", "Mfg Completed"):
            has_incomplete = any(
                not processes.get(k, {}).get("completionDate")
                for k in ordered_keys
            )
            if has_incomplete:
                alerts.append({
                    "type": "STALE_WIP",
                    "severity": "MEDIUM",
                    "process": None,
                    "message": f"Work-in-progress with no update for {(now - target_date).days} days past target",
                    "targetDate": target_date.isoformat(),
                })

    # 4. Missing Date — target exists, process completion blank, target is past
    if target_date and now > target_date:
        for proc_key in ordered_keys:
            proc_info = processes.get(proc_key, {})
            comp_date = proc_info.get("completionDate")
            if not comp_date:
                alerts.append({
                    "type": "MISSING_DATE",
                    "severity": "LOW",
                    "process": proc_key,
                    "message": f"{proc_info.get('operationName', proc_key)} has no completion date — target was {target_date.strftime('%d-%b-%Y')}",
                    "targetDate": target_date.isoformat(),
                })

    return alerts


# ─── Excel Reading ─────────────────────────────────────────────────────────────

def _build_col_index(header_row):
    """Build a column name -> index mapping from a header row."""
    mapping = {}
    for idx, cell_val in enumerate(header_row):
        if cell_val:
            key = str(cell_val).strip()
            mapping[key] = idx
    return mapping


def _safe_get(row, col_map, col_name, default=None):
    """Safely get a value from a row using column mapping."""
    idx = col_map.get(col_name)
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    return val if val is not None else default


def _read_all_pms_data():
    """
    Read ALL sheets from PMS.xlsx and merge into a unified dataset.
    Returns a list of records with all process dates and risk data.
    """
    if not os.path.exists(PMS_FILE_PATH):
        print(f"PMS file not found: {PMS_FILE_PATH}")
        return []

    try:
        wb = openpyxl.load_workbook(PMS_FILE_PATH, data_only=True)
    except Exception as e:
        print(f"Error opening PMS file: {e}")
        return []

    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue

        # Find header row
        header_row = None
        header_row_idx = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True)):
            row_vals = [str(v).strip() if v else "" for v in row]
            # Detect header by looking for key columns
            if "Station" in row_vals or "W/O No" in row_vals or "Part No" in row_vals:
                header_row = row
                header_row_idx = row_idx + 1
                break

        if not header_row:
            continue

        col_map = _build_col_index(header_row)

        # Check if this sheet has a KO No column (BOM Tracker has it, detail sheets don't)
        has_ko = "KO No" in col_map

        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
            if not row or not any(row):
                continue

            # Get Work Order number
            wo_no = _safe_get(row, col_map, "W/O No")
            if not wo_no or not isinstance(wo_no, str) or not wo_no.strip().startswith("WO/"):
                continue

            wo_no = wo_no.strip()

            # Basic fields
            station = _safe_get(row, col_map, "Station", "")
            build = _safe_get(row, col_map, "Build", "")
            prog_no = _safe_get(row, col_map, "Prog No", "")
            proj_type = _safe_get(row, col_map, "Proj_Type", "")
            ko_date = _parse_date(_safe_get(row, col_map, "Mfg KO Date"))
            part_no = _safe_get(row, col_map, "Part No", "")
            bom_qty = _safe_get(row, col_map, "BOM Qty", 0)
            qpl = _safe_get(row, col_map, "QPL", 0)
            mfg_qty = _safe_get(row, col_map, "Mfg Qty", 0)
            rm_grade = _safe_get(row, col_map, "R/M Grade", "")
            location = _safe_get(row, col_map, "Location", "")
            mc_category = _safe_get(row, col_map, "Mc Category", "")
            complexity = _safe_get(row, col_map, "Complexity", "")
            dri = _safe_get(row, col_map, "DRI", "")
            target_date = _parse_date(_safe_get(row, col_map, "Mc Target_Date"))
            present_status = _safe_get(row, col_map, "Present Status", "")
            programmer = _safe_get(row, col_map, "Programmer Name", "")
            cam_shift = _safe_get(row, col_map, "CAM Shift", "")
            cam_comp_date = _parse_date(_safe_get(row, col_map, "CAM Comp_Date"))
            rdc_no = _safe_get(row, col_map, "RDC No", "")
            status = _safe_get(row, col_map, "Status", "Open")
            remarks = _safe_get(row, col_map, "Remarks", "")
            ko_number = _safe_get(row, col_map, "KO No", "") if has_ko else ""

            # Process dates
            processes = {}
            for proc_key, col_info in _PROCESS_DATE_COLS.items():
                comp_date = _parse_date(_safe_get(row, col_map, col_info["date_col"]))
                comp_qty = _safe_get(row, col_map, col_info["qty_col"], 0)
                # Find the operation name from STANDARD_PROCESSES
                op_name = next((p["operationName"] for p in STANDARD_PROCESSES if p["processKey"] == proc_key), proc_key)
                processes[proc_key] = {
                    "operationName": op_name,
                    "completionDate": comp_date,
                    "completionQty": _safe_float(comp_qty),
                }

            # Compute risk
            risk_level, risk_score = calculate_risk_level(target_date, complexity)

            record = {
                "workOrderNumber": wo_no,
                "koNumber": str(ko_number).strip() if ko_number else "",
                "station": str(station) if station else "",
                "jobName": str(station) if station else "",
                "build": str(build) if build else "",
                "progNo": str(prog_no) if prog_no else "",
                "projType": str(proj_type) if proj_type else "",
                "koDate": ko_date.isoformat() if ko_date else None,
                "partNumber": str(part_no) if part_no else "",
                "bomQty": _safe_float(bom_qty),
                "qpl": _safe_float(qpl),
                "mfgQty": _safe_float(mfg_qty),
                "batchQuantity": int(_safe_float(mfg_qty)) if mfg_qty else 1,
                "rmGrade": str(rm_grade) if rm_grade else "",
                "location": str(location) if location else "",
                "mcCategory": str(mc_category) if mc_category else "",
                "workCenter": str(mc_category) if mc_category else "",
                "complexity": str(complexity) if complexity else "",
                "dri": str(dri) if dri else "",
                "targetDate": target_date,
                "targetDateStr": target_date.strftime("%Y-%m-%d") if target_date else None,
                "presentStatus": str(present_status) if present_status else "",
                "programmer": str(programmer) if programmer else "",
                "camShift": str(cam_shift) if cam_shift else "",
                "camCompDate": cam_comp_date.isoformat() if cam_comp_date else None,
                "rdcNo": str(rdc_no) if rdc_no else "",
                "status": str(status) if status else "Open",
                "remarks": str(remarks) if remarks else "",
                "partRevision": str(build) if build else "A",
                "riskLevel": risk_level,
                "riskScore": risk_score,
                "processes": processes,
                "sourceSheet": sheet_name,
            }

            # Detect alerts
            record["alerts"] = detect_date_mismatches(record)
            record["alertCount"] = len(record["alerts"])

            records.append(record)

    return records


def _safe_float(val):
    """Safely convert to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ─── Serialization Helper ─────────────────────────────────────────────────────

def _serialize_record(record):
    """Convert a record to JSON-serializable dict (handle datetime objects)."""
    r = dict(record)
    # Remove raw datetime target_date, keep targetDateStr
    if "targetDate" in r and isinstance(r["targetDate"], datetime):
        r["targetDate"] = r["targetDate"].isoformat()
    # Serialize process dates
    procs = {}
    for pkey, pval in r.get("processes", {}).items():
        p = dict(pval)
        if "completionDate" in p and isinstance(p["completionDate"], datetime):
            p["completionDate"] = p["completionDate"].isoformat()
        procs[pkey] = p
    r["processes"] = procs
    return r


# ─── Public API Functions ─────────────────────────────────────────────────────

_cache = {"data": None, "timestamp": None}
_CACHE_TTL = 60  # seconds
_cache_lock = threading.Lock()
_bg_thread_started = False

def _background_refresh():
    while True:
        try:
            # Sleep first if data is already loaded so we don't immediately refresh
            if _cache["data"] is not None:
                time.sleep(_CACHE_TTL)
            
            data = _read_all_pms_data()
            with _cache_lock:
                _cache["data"] = data
                _cache["timestamp"] = datetime.now()
                
            if _cache["data"] is None:
                time.sleep(_CACHE_TTL)
        except Exception as e:
            print(f"Background refresh failed: {e}")
            time.sleep(10)

def _get_cached_data():
    """Get PMS data instantly (cache is refreshed in background thread)."""
    global _bg_thread_started
    
    if not _bg_thread_started:
        with _cache_lock:
            if not _bg_thread_started:
                if _cache["data"] is None:
                    _cache["data"] = _read_all_pms_data()
                    _cache["timestamp"] = datetime.now()
                t = threading.Thread(target=_background_refresh, daemon=True)
                t.start()
                _bg_thread_started = True

    return _cache["data"] or []


def get_all_work_orders():
    """Return list of all work orders with risk levels for listing/dropdown."""
    records = _get_cached_data()
    # Deduplicate by WO number, keep first occurrence
    seen = set()
    result = []
    for r in records:
        wo = r["workOrderNumber"]
        if wo not in seen:
            seen.add(wo)
            serialized = _serialize_record(r)
            result.append({
                "workOrderNumber": serialized["workOrderNumber"],
                "koNumber": serialized["koNumber"],
                "jobName": serialized["jobName"],
                "partNumber": serialized["partNumber"],
                "partRevision": serialized["partRevision"],
                "batchQuantity": serialized["batchQuantity"],
                "riskLevel": serialized["riskLevel"],
                "riskScore": serialized["riskScore"],
                "targetDate": serialized.get("targetDateStr"),
                "presentStatus": serialized["presentStatus"],
                "status": serialized["status"],
                "complexity": serialized["complexity"],
                "alertCount": serialized["alertCount"],
                "sourceSheet": serialized["sourceSheet"],
            })
    return result


def get_work_order_details(wo_number: str):
    """Return full details for a specific WO, including auto-generated 7-step process pipeline."""
    records = _get_cached_data()
    for r in records:
        if r["workOrderNumber"] == wo_number:
            serialized = _serialize_record(r)

            # Build the 7 standard process steps with completion data from PMS
            steps = []
            for proc in STANDARD_PROCESSES:
                proc_data = serialized.get("processes", {}).get(proc["processKey"], {})
                step = {
                    "stepNumber": proc["stepNumber"],
                    "processKey": proc["processKey"],
                    "operationName": proc["operationName"],
                    "workCenter": proc.get("workCenter", "WIP"),
                    "instructions": proc["instructions"],
                    "requiredSop": proc["requiredSop"],
                    "completionDate": proc_data.get("completionDate"),
                    "completionQty": proc_data.get("completionQty", 0),
                    "status": "Completed" if proc_data.get("completionDate") else "Pending",
                }
                # Special Process: mark N/A if no SP date and no RDC number
                if proc["processKey"] == "special_process":
                    if not proc_data.get("completionDate") and not serialized.get("rdcNo"):
                        step["status"] = "N/A"
                steps.append(step)

            return {
                **serialized,
                "steps": steps,
                "alerts": serialized.get("alerts", []),
            }
    return None


def get_ko_numbers():
    """Return list of unique KO numbers for dropdown/search."""
    records = _get_cached_data()
    ko_set = set()
    result = []
    for r in records:
        ko = r.get("koNumber", "").strip()
        if ko and ko not in ko_set:
            ko_set.add(ko)
            result.append(ko)
    return sorted(result)


def get_work_orders_by_ko(ko_number: str):
    """Return all work orders grouped under a specific KO number."""
    records = _get_cached_data()
    result = []
    for r in records:
        if r.get("koNumber", "").strip() == ko_number.strip():
            result.append(_serialize_record(r))
    return result


def get_risk_summary():
    """Return dashboard summary: counts by risk level + total alerts."""
    records = _get_cached_data()
    summary = {"high": 0, "medium": 0, "low": 0, "total": len(records), "totalAlerts": 0}
    alert_breakdown = {"OVERDUE_PROCESS": 0, "SEQUENCE_VIOLATION": 0, "STALE_WIP": 0, "MISSING_DATE": 0}

    for r in records:
        risk = r.get("riskLevel", "LOW")
        if risk == "HIGH":
            summary["high"] += 1
        elif risk == "MEDIUM":
            summary["medium"] += 1
        else:
            summary["low"] += 1

        for alert in r.get("alerts", []):
            summary["totalAlerts"] += 1
            atype = alert.get("type", "")
            if atype in alert_breakdown:
                alert_breakdown[atype] += 1

    summary["alertBreakdown"] = alert_breakdown
    return summary


def get_alerts(limit=50):
    """Return all active date mismatch alerts across all work orders, sorted by severity."""
    records = _get_cached_data()
    all_alerts = []
    severity_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}

    for r in records:
        for alert in r.get("alerts", []):
            all_alerts.append({
                **alert,
                "workOrderNumber": r["workOrderNumber"],
                "koNumber": r.get("koNumber", ""),
                "partNumber": r.get("partNumber", ""),
                "jobName": r.get("jobName", ""),
            })

    all_alerts.sort(key=lambda a: severity_order.get(a.get("severity", "LOW"), 2))
    return all_alerts[:limit]


def invalidate_cache():
    """Force re-read of PMS data in the background (returns instantly)."""
    with _cache_lock:
        _cache["data"] = None
        _cache["timestamp"] = None
