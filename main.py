"""
Production Management System — FastAPI Backend
Enhanced with 7-step process pipeline, risk classification, KO number integration,
and date mismatch alert detection.
"""

import sys
import os

# Ensure local imports resolve correctly
sys.path.insert(0, os.path.dirname(__file__))

from fastapi import FastAPI, HTTPException, Response, Request, UploadFile, File, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from typing import List, Optional
import tempfile
import shutil

from models import (
    SignupRequest, LoginWithPasswordRequest, GoogleLoginRequest, ForgotPasswordRequest,
    VerifyOTPRequest, ResetPasswordRequest, LoginResponse, UserRole,
    CreateCardRequest, SignOffRequest, DeviationRequest, ResolveRequest,
    IQCFailRequest, IQCReinspectRequest,
    UpdatePreferencesRequest, SimulateNotifRequest, BulkCreateRequest
)
from auth import create_access_token, verify_google_token, get_current_user, require_role
from cache import redis_cache
from database import db
import pms_db

app = FastAPI(
    title="ASM Production Management System API",
    description="Enterprise manufacturing route card & process tracking with risk classification",
    version="3.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Auth ──────────────────────────────────────────────────────────────────────

@app.post("/api/auth/signup")
def signup(req: SignupRequest):
    user = db.register_user(req.email.strip(), req.name.strip(), req.role.value if req.role else "Operator", req.password)
    if not user:
        raise HTTPException(400, "Email already registered.")
    db.log_activity(user["id"], user["name"], user["email"], "user signup", f"User {user['email']} signed up as {user['role']}")
    return {"success": True, "user": {k:v for k,v in user.items() if k != "password"}}

@app.post("/api/auth/login")
def login(req: LoginWithPasswordRequest):
    user = db.authenticate_user(req.email.strip(), req.password)
    if not user:
        raise HTTPException(401, "Invalid email or password.")
    
    # Generate JWT
    token = create_access_token(user["id"], user["role"])
    
    db.log_activity(user["id"], user["name"], user["email"], "user login", f"User successfully authenticated")
    return {"success": True, "token": token, "user": {k:v for k,v in user.items() if k != "password"}}

@app.post("/api/auth/google")
def google_login(req: GoogleLoginRequest):
    idinfo = verify_google_token(req.idToken)
    email = idinfo.get("email")
    name = idinfo.get("name")
    
    # Check if user exists
    user = next((u for u in db.users if u["email"] == email), None)
    if not user:
        # Auto-register google users with requested role or default to Operator
        role_value = req.role.value if req.role else "Operator"
        user = db.register_user(email, name, role_value, "google-sso-placeholder")
    
    token = create_access_token(user["id"], user["role"])
    db.log_activity(user["id"], user["name"], user["email"], "google login", f"User successfully authenticated via Google")
    return {"success": True, "token": token, "user": {k:v for k,v in user.items() if k != "password"}}

@app.post("/api/auth/forgot-password")
def forgot_password(req: ForgotPasswordRequest):
    otp = db.generate_otp(req.email.strip())
    if not otp:
        raise HTTPException(404, "Email not found.")
    
    # Mock sending email by printing to terminal
    print("\n" + "="*50)
    print(f"📧 MOCK EMAIL SENT TO: {req.email.strip()}")
    print(f"🔒 YOUR OTP IS: {otp}")
    print("="*50 + "\n")
    
    return {"success": True, "message": "OTP sent to email."}

@app.post("/api/auth/verify-otp")
def verify_otp(req: VerifyOTPRequest):
    if not db.verify_otp(req.email.strip(), req.otp.strip()):
        raise HTTPException(400, "Invalid or expired OTP.")
    return {"success": True}

@app.post("/api/auth/reset-password")
def reset_password(req: ResetPasswordRequest):
    if not db.reset_password(req.email.strip(), req.otp.strip(), req.newPassword):
        raise HTTPException(400, "Invalid OTP or Reset Failed.")
    return {"success": True, "message": "Password reset successfully."}


# ─── PMS Endpoints ─────────────────────────────────────────────────────────────

@app.get("/api/pms/work-orders")
@redis_cache(ttl=60)
def get_pms_work_orders():
    """List all work orders from PMS Excel with risk levels and alert counts."""
    wos = pms_db.get_all_work_orders()
    return wos

@app.get("/api/pms/work-orders/{wo_number:path}")
def get_pms_work_order_details(wo_number: str):
    """Get full details for a specific work order including 7-step process pipeline."""
    details = pms_db.get_work_order_details(wo_number)
    if not details:
        raise HTTPException(status_code=404, detail="Work Order not found in PMS")
    return details

@app.get("/api/pms/ko-numbers")
def get_ko_numbers():
    """List all unique KO numbers for dropdown/autocomplete."""
    return pms_db.get_ko_numbers()

@app.get("/api/pms/work-orders-by-ko/{ko_number}")
def get_work_orders_by_ko(ko_number: str):
    """Get all work orders grouped under a specific KO number."""
    wos = pms_db.get_work_orders_by_ko(ko_number)
    if not wos:
        raise HTTPException(404, f"No work orders found for KO number: {ko_number}")
    return wos

@app.get("/api/pms/risk-summary")
@redis_cache(ttl=60)
def get_risk_summary():
    """Dashboard risk summary — counts by HIGH/MEDIUM/LOW + alert breakdown."""
    return pms_db.get_risk_summary()

@app.get("/api/pms/alerts")
@redis_cache(ttl=60)
def get_alerts(limit: int = 50):
    """Get all active date mismatch alerts, sorted by severity."""
    return pms_db.get_alerts(limit=limit)

@app.post("/api/pms/upload-bom")
async def upload_bom(file: UploadFile = File(...)):
    """Upload a BOM Excel file to extract KO numbers and work order data."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only Excel files (.xlsx, .xls) are accepted.")
    
    # Save uploaded file temporarily
    try:
        tmp_dir = tempfile.mkdtemp()
        tmp_path = os.path.join(tmp_dir, file.filename)
        with open(tmp_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Parse the uploaded BOM using openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        
        ko_numbers = set()
        work_orders = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                continue
            
            # Find header
            header_row = None
            header_idx = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True)):
                row_vals = [str(v).strip() if v else "" for v in row]
                if "Station" in row_vals or "W/O No" in row_vals:
                    header_row = row
                    header_idx = row_idx + 1
                    break
            
            if not header_row:
                continue
            
            col_map = {}
            for idx, val in enumerate(header_row):
                if val:
                    col_map[str(val).strip()] = idx
            
            for row in ws.iter_rows(min_row=header_idx + 1, max_row=ws.max_row, values_only=True):
                if not row or not any(row):
                    continue
                
                # Extract KO Number
                ko_idx = col_map.get("KO No")
                if ko_idx is not None and ko_idx < len(row) and row[ko_idx]:
                    ko_numbers.add(str(row[ko_idx]).strip())
                
                # Extract WO Number
                wo_idx = col_map.get("W/O No")
                if wo_idx is not None and wo_idx < len(row) and row[wo_idx]:
                    wo = str(row[wo_idx]).strip()
                    if wo.startswith("WO/"):
                        part_idx = col_map.get("Part No")
                        station_idx = col_map.get("Station")
                        work_orders.append({
                            "workOrderNumber": wo,
                            "partNumber": str(row[part_idx]).strip() if part_idx and part_idx < len(row) and row[part_idx] else "",
                            "station": str(row[station_idx]).strip() if station_idx and station_idx < len(row) and row[station_idx] else "",
                        })
        
        # Cleanup
        shutil.rmtree(tmp_dir, ignore_errors=True)
        
        return {
            "success": True,
            "fileName": file.filename,
            "sheetsProcessed": len(wb.sheetnames),
            "koNumbers": sorted(list(ko_numbers)),
            "workOrderCount": len(work_orders),
            "workOrders": work_orders[:100],  # Limit to first 100
        }
    except Exception as e:
        raise HTTPException(500, f"Error processing BOM file: {str(e)}")

@app.post("/api/bom/parse")
async def parse_bom(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only Excel files (.xlsx, .xls) are accepted.")
    
    try:
        tmp_dir = tempfile.mkdtemp()
        tmp_path = os.path.join(tmp_dir, file.filename)
        with open(tmp_path, "wb") as f:
            content = await file.read()
            f.write(content)
            
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        
        ko_number = ""
        items = []
        
        # Scan ALL sheets (BOM data may not be on the active/first sheet)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                continue
            
            # --- Find KO / Program Number in the first 10 rows ---
            if not ko_number:
                for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True):
                    row_vals = [str(v).strip() if v is not None else "" for v in row]
                    for search_key in ("Program No", "Program No:", "KO No", "KO Number"):
                        for idx, v in enumerate(row_vals):
                            if search_key in v:
                                for next_idx in range(idx + 1, len(row_vals)):
                                    if row_vals[next_idx] and row_vals[next_idx] != "-":
                                        ko_number = row_vals[next_idx]
                                        break
                                break
                    if ko_number:
                        break
            
            # --- Find header row containing Part No / Drawing Number + Description ---
            header_row_idx = 0
            col_map = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True)):
                row_vals = [str(v).strip() if v is not None else "" for v in row]
                # Accept header if it has Description AND any part identifier column
                has_desc = "Description" in row_vals
                has_part = "Part No" in row_vals or "Drawing Number" in row_vals or "Part Number" in row_vals
                if has_desc and has_part:
                    header_row_idx = row_idx + 1
                    for i, val in enumerate(row_vals):
                        if val:
                            col_map[val] = i
                    break
            
            if not header_row_idx or not col_map:
                continue  # Try next sheet
            
            # --- Determine which column to use for Part No ---
            # Priority: "Drawing Number" > "Part No" > "Part Number" 
            # (in ASM BOMs, "Part No" is often "-" while "Drawing Number" has the real identifier)
            part_col_key = None
            for candidate in ("Drawing Number", "Part No", "Part Number"):
                if candidate in col_map:
                    part_col_key = candidate
                    break
            
            desc_col = col_map.get("Description")
            qty_col = col_map.get("Qty", col_map.get("Quantity"))
            
            if part_col_key is None or desc_col is None:
                continue  # Try next sheet
            
            part_col = col_map[part_col_key]
            
            # --- Parse data rows ---
            for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                row_vals = [str(v).strip() if v is not None else "" for v in row]
                
                part_no = row_vals[part_col] if part_col < len(row_vals) else ""
                desc = row_vals[desc_col] if desc_col < len(row_vals) else ""
                qty = row_vals[qty_col] if qty_col is not None and qty_col < len(row_vals) else ""
                
                # If Drawing Number is "-", try falling back to Part No column
                if (not part_no or part_no == "-") and part_col_key == "Drawing Number" and "Part No" in col_map:
                    fallback = row_vals[col_map["Part No"]] if col_map["Part No"] < len(row_vals) else ""
                    if fallback and fallback != "-" and fallback != "None":
                        part_no = fallback
                
                if part_no and part_no != "-" and part_no != "None":
                    items.append({
                        "partNo": part_no,
                        "description": desc,
                        "qty": qty
                    })
            
            # If we found items on this sheet, don't scan further sheets
            if items:
                break
                    
        shutil.rmtree(tmp_dir, ignore_errors=True)
        
        if not ko_number:
            ko_number = "UNKNOWN_KO"
        
        return {
            "success": True,
            "koNumber": ko_number,
            "items": items
        }
    except Exception as e:
        raise HTTPException(500, f"Error parsing BOM file: {str(e)}")

@app.post("/api/bom/bulk-create")
def bulk_create_bom_cards(req: BulkCreateRequest, user: dict = Depends(require_role([UserRole.Admin, UserRole.ShiftEngineer]))):
    from datetime import datetime, timedelta
    created_count = 0
    for item in req.items:
        part_no = item.get("partNo", "")
        if not part_no:
            continue
            
        wo_number = f"{req.koNumber}-{req.bomNumber}-{part_no}"
        
        # Check if exists
        exists = False
        for c in db.cards:
            if c["workOrderNumber"] == wo_number:
                exists = True
                break
        if exists:
            continue
        
        card_data = {
            "workOrderNumber": wo_number,
            "koNumber": req.koNumber,
            "bomNumber": req.bomNumber,
            "partNumber": part_no,
            "jobName": item.get("description", ""),
            "targetDate": (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"),
            "riskLevel": "LOW",
            "complexity": "Medium"
        }
        
        db.create_route_card(card_data)
        created_count += 1
        
    return {"success": True, "createdCount": created_count, "message": f"Successfully created {created_count} route cards."}


@app.post("/api/pms/refresh")
def refresh_pms_cache():
    """Force re-read of PMS data from Excel."""
    pms_db.invalidate_cache()
    return {"success": True, "message": "PMS cache invalidated. Data will be re-read on next request."}


# ─── Route Cards ───────────────────────────────────────────────────────────────

@app.get("/api/route-cards")
def get_all_cards():
    return db.get_all_route_cards()


@app.get("/api/route-cards/{card_id}")
def get_card(card_id: str):
    card = db.get_route_card_details(card_id)
    if not card:
        raise HTTPException(404, "Route card not found.")
    return card


@app.get("/api/route-cards/ko/{ko_number}")
def get_cards_by_ko(ko_number: str):
    """Get all route cards grouped under a specific KO number."""
    cards = db.get_route_cards_by_ko(ko_number)
    return cards


@app.post("/api/route-cards", status_code=210)
def create_card(req: CreateCardRequest, current_user: dict = Depends(require_role([UserRole.Admin, UserRole.ShiftEngineer]))):
    if not req.jobName or not req.partNumber or not req.batchQuantity or not req.workOrderNumber:
        raise HTTPException(400, "Missing required fields.")
    if not req.koNumber:
        raise HTTPException(400, "KO Number is required. All parts must be grouped under a KO number.")

    card_data = {
        "jobName": req.jobName,
        "partNumber": req.partNumber,
        "partRevision": req.partRevision or "A",
        "batchQuantity": int(req.batchQuantity),
        "workOrderNumber": req.workOrderNumber,
        "koNumber": req.koNumber,
        "riskLevel": req.riskLevel or "LOW",
        "riskScore": {"HIGH": 3, "MEDIUM": 2, "LOW": 1}.get(req.riskLevel, 1) if req.riskLevel else 1,
        "complexity": req.complexity or "",
        "targetDate": req.targetDate or "",
        "notes": req.notes or "",
        "createdBy": req.createdBy or "System Creator",
    }

    # If steps are provided, use them; otherwise auto-generate 7 standard processes
    steps_data = None
    if req.steps:
        steps_data = [
            {
                "stepNumber": s.stepNumber or (i + 1) * 10,
                "processKey": s.processKey or "",
                "operationName": s.operationName,
                "workCenter": s.workCenter or "WIP",
                "instructions": s.instructions,
                "requiredSop": s.requiredSop or "SOP-GEN-01",
            }
            for i, s in enumerate(req.steps)
        ]

    new_card = db.create_route_card(card_data, steps_data)

    author_id = req.userId or "u-3"
    author_name = req.createdBy or "Sarah Engineer"
    author_email = req.userEmail or "engineer1@asmltd.com"
    db.log_activity(author_id, author_name, author_email, "route card creation",
                    f"Created route card {new_card['cardNumber']} ({new_card['jobName']}) under KO: {req.koNumber}")

    for uid in ["u-1", "u-2", "u-3"]:
        db.create_notification(uid, "New Route Card Assigned",
            f"New Route Card {new_card['cardNumber']} [WO: {new_card['workOrderNumber']}] [KO: {req.koNumber}] — Risk: {card_data['riskLevel']}",
            "New Route Card Assigned")

    return new_card


# ─── Step Actions ──────────────────────────────────────────────────────────────

@app.put("/api/route-cards/{card_id}/steps/{step_id}/sign-off")
def sign_off_step(card_id: str, step_id: str, req: SignOffRequest, current_user: dict = Depends(get_current_user)):
    if not req.operatorName or not req.operatorRole:
        raise HTTPException(400, "Sign-off requires operator name and role context.")
        
    card = db.get_route_card_details(card_id)
    if not card:
        raise HTTPException(404, "Route Card not found.")
        
    target_step = next((s for s in card["steps"] if s["id"] == step_id), None)
    if not target_step:
        raise HTTPException(404, "Step not found.")
        
    # Enforce RBAC
    role = current_user.get("role")
    wc = current_user.get("workCenter", "")
    if role == "Operator" and wc and wc.lower() != target_step.get("processKey", "").lower():
        raise HTTPException(403, f"Access denied. You are assigned to '{wc}' but this step requires '{target_step.get('processKey')}'.")

    updated = db.update_step_sign_off(
        card_id, step_id, req.operatorName, req.operatorRole.value, req.remarks, req.completionQty
    )

    db.log_activity(
        req.userId or current_user.get("id"), req.operatorName, req.userEmail or current_user.get("email"),
        "step sign-off",
        f"Signed off {target_step['operationName']} (Step {target_step['stepNumber']}) on {updated['cardNumber']}"
    )

    if updated["status"] == "Completed":
        for uid in ["u-1", "u-3"]:
            db.create_notification(uid, "Admin Action Required",
                f"Route Card {updated['cardNumber']} [{updated['workOrderNumber']}] completed all process steps. Verification stamp pending.",
                "Admin Action Required")

    return updated


@app.put("/api/route-cards/{card_id}/steps/{step_id}/progress")
def progress_step(card_id: str, step_id: str, current_user: dict = Depends(get_current_user)):
    card = db.get_route_card_details(card_id)
    target_step = next((s for s in card["steps"] if s["id"] == step_id), None) if card else None
    
    role = current_user.get("role")
    wc = current_user.get("workCenter", "")
    if target_step and role == "Operator" and wc and wc.lower() != target_step.get("processKey", "").lower():
        raise HTTPException(403, "Access denied. Process mismatch.")
        
    updated = db.update_step_progress(card_id, step_id)
    if not updated:
        raise HTTPException(404, "Route Card or Step not found.")
    return updated


@app.put("/api/route-cards/{card_id}/steps/{step_id}/flag")
def flag_step(card_id: str, step_id: str, req: DeviationRequest, current_user: dict = Depends(get_current_user)):
    if not req.reason:
        raise HTTPException(400, "Flagging a deviation requires a reason.")

    updated = db.update_step_deviation(card_id, step_id, req.reason, req.remarks)
    if not updated:
        raise HTTPException(404, "Route Card or Step not found.")

    target_step = next((s for s in updated["steps"] if s["id"] == step_id), None)
    db.log_activity(
        req.userId or current_user.get("id"), req.operatorName or current_user.get("name"), req.userEmail or current_user.get("email"),
        "deviation flagged",
        f"Flagged deviation on {target_step['operationName'] if target_step else '?'} for {updated['cardNumber']}. Reason: {req.reason}"
    )

    for uid in ["u-1", "u-3"]:
        db.create_notification(uid, "Admin Action Required",
            f"DEVIATION on {updated['cardNumber']}: {target_step['operationName'] if target_step else '?'} deviated. Reason: {req.reason[:100]}",
            "Admin Action Required")

    return updated


@app.put("/api/route-cards/{card_id}/steps/{step_id}/iqc-fail")
def iqc_fail(card_id: str, step_id: str, req: IQCFailRequest, current_user: dict = Depends(get_current_user)):
    """Mark IQC as failed — reject / return to vendor."""
    if not req.reason:
        raise HTTPException(400, "IQC failure requires a reason.")
        
    role = current_user.get("role")
    wc = current_user.get("workCenter", "")
    if role == "Operator" and wc and wc.lower() != "iqc":
        raise HTTPException(403, "Access denied. Only IQC operators can perform this action.")

    updated = db.update_iqc_fail(card_id, step_id, req.reason, req.remarks)
    if not updated:
        raise HTTPException(404, "Route Card, Step not found, or step is not IQC.")

    db.log_activity(
        req.userId or current_user.get("id"), req.operatorName or current_user.get("name"), req.userEmail or current_user.get("email"),
        "IQC failed",
        f"IQC FAILED on {updated['cardNumber']}: {req.reason}"
    )

    for uid in ["u-1", "u-3"]:
        db.create_notification(uid, "IQC Failed",
            f"🔴 IQC FAILED on {updated['cardNumber']} [{updated['workOrderNumber']}]: {req.reason[:150]}. Material rejected / returned to vendor.",
            "IQC Failed")

    return updated


@app.put("/api/route-cards/{card_id}/steps/{step_id}/iqc-reinspect")
def iqc_reinspect(card_id: str, step_id: str, req: IQCReinspectRequest, current_user: dict = Depends(get_current_user)):
    """Re-inspect after vendor return — resets IQC step for re-inspection."""
    role = current_user.get("role")
    wc = current_user.get("workCenter", "")
    if role == "Operator" and wc and wc.lower() != "iqc":
        raise HTTPException(403, "Access denied. Only IQC operators can perform this action.")
        
    updated = db.update_iqc_reinspect(card_id, step_id, req.remarks)
    if not updated:
        raise HTTPException(404, "Route Card, Step not found, or step is not IQC.")

    db.log_activity(
        req.userId or current_user.get("id"), req.operatorName or current_user.get("name"), req.userEmail or current_user.get("email"),
        "IQC re-inspection",
        f"IQC re-inspection initiated on {updated['cardNumber']} after vendor return."
    )

    for uid in ["u-1", "u-3"]:
        db.create_notification(uid, "Admin Action Required",
            f"IQC re-inspection initiated on {updated['cardNumber']} [{updated['workOrderNumber']}]. Vendor material returned for re-check.",
            "Admin Action Required")

    return updated


@app.put("/api/route-cards/{card_id}/steps/{step_id}/resolve")
def resolve_step(card_id: str, step_id: str, req: ResolveRequest, current_user: dict = Depends(require_role([UserRole.Admin, UserRole.ShiftEngineer]))):
    if not req.remarks or not req.engineerName:
        raise HTTPException(400, "Resolution requires engineer name and remarks.")

    updated = db.resolve_deviation(card_id, step_id, req.remarks, req.engineerName)
    if not updated:
        raise HTTPException(404, "Route Card or Step not found.")

    db.log_activity(
        req.userId or current_user.get("id"), req.engineerName, current_user.get("email"),
        "deviation resolved",
        f"Resolved deviation on {updated['cardNumber']}: {req.remarks[:100]}"
    )
    return updated


# ─── Export ────────────────────────────────────────────────────────────────

@app.get("/api/route-cards/{card_id}/export")
def export_excel(card_id: str, userId: Optional[str] = None, userName: Optional[str] = None, userEmail: Optional[str] = None):
    from pdf_gen import generate_traveler_pdf
    from datetime import datetime
    
    card = db.get_route_card_details(card_id)
    if not card:
        raise HTTPException(404, "Route card not found.")
        
    try:
        pdf_bytes = generate_traveler_pdf(card)
    except Exception as e:
        print(f"Error generating PDF: {e}")
        raise HTTPException(500, "Error generating PDF.")
        
    if userId and userName:
        db.log_activity(userId, userName, userEmail or "system@asmltd.com", "PDF generation",
                        f"Exported route card {card['cardNumber']} to PDF")
                        
    date_str = datetime.now().strftime("%Y%m%d")
    wo = card.get("workOrderNumber", "UNKNOWN").replace("/", "_")
    filename = f"WO-{wo}-{date_str}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Users / Notifications ─────────────────────────────────────────────────────

@app.get("/api/users")
def get_users():
    return db.get_all_users()


@app.get("/api/users/{user_id}/notifications")
def get_notifications(user_id: str):
    return db.get_notifications(user_id)


@app.put("/api/users/{user_id}/notifications/{notif_id}/read")
def mark_read(user_id: str, notif_id: str):
    ok = db.mark_notification_read(notif_id)
    return {"success": ok}


@app.put("/api/users/{user_id}/notifications/read-all")
def mark_all_read(user_id: str):
    ok = db.mark_all_notifications_read(user_id)
    return {"success": ok}


@app.post("/api/users/{user_id}/notifications/simulate")
def simulate_notification(user_id: str, req: SimulateNotifRequest):
    n = db.create_notification(user_id, req.title, req.message, req.type.value)
    return n or {"error": "Suppressed by preferences"}


@app.get("/api/users/{user_id}/preferences")
def get_preferences(user_id: str):
    return db.get_user_preferences(user_id)


@app.put("/api/users/{user_id}/preferences")
def update_preferences(user_id: str, req: UpdatePreferencesRequest):
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    return db.update_user_preferences(user_id, updates)


# ─── Activity Logs ─────────────────────────────────────────────────────────────

@app.get("/api/activity-logs")
def get_activity_logs():
    return db.get_activity_logs()


# ─── Admin ────────────────────────────────────────────────────────────────────

@app.post("/api/admin/reset")
def reset_database():
    db.reset()
    pms_db.invalidate_cache()
    return {"success": True, "message": "Database reset to factory defaults."}


# ─── Health ───────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    return {"status": "ok", "service": "ASM Production Management System", "version": "3.0.0"}


# ─── Entry Point ──────────────────────────────────────────────────────────────

flutter_build_dir = os.path.join(os.path.dirname(__file__), "flutter_app", "build", "web")

@app.middleware("http")
async def spa_middleware(request: Request, call_next):
    response = await call_next(request)
    if response.status_code == 404 and not request.url.path.startswith("/api/"):
        index_path = os.path.join(flutter_build_dir, "index.html")
        if os.path.exists(index_path):
            return FileResponse(index_path)
    return response

os.makedirs(flutter_build_dir, exist_ok=True)
app.mount("/", StaticFiles(directory=flutter_build_dir, html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
