import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\A30061\Downloads\V6Y_MCPL_MLB (EVT BUILD)_V2_10-07-2026.xlsx', data_only=True)
ws = wb.active

ko_number = ""
for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
    row_vals = [str(v).strip() if v is not None else "" for v in row]
    if any("Program No" in v for v in row_vals):
        for idx, v in enumerate(row_vals):
            if "Program No" in v:
                for next_idx in range(idx + 1, len(row_vals)):
                    if row_vals[next_idx]:
                        ko_number = row_vals[next_idx]
                        break
                break

header_row_idx = 0
col_map = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
    row_vals = [str(v).strip() if v is not None else "" for v in row]
    if "Part No" in row_vals and "Description" in row_vals:
        header_row_idx = row_idx + 1
        for i, val in enumerate(row_vals):
            if val:
                col_map[val] = i
        break

items = []
for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
    if not row or all(v is None for v in row):
        continue
    row_vals = [str(v).strip() if v is not None else "" for v in row]
    part_no = row_vals[col_map["Part No"]] if "Part No" in col_map and col_map["Part No"] < len(row_vals) else ""
    desc = row_vals[col_map["Description"]] if "Description" in col_map and col_map["Description"] < len(row_vals) else ""
    qty = row_vals[col_map["Qty"]] if "Qty" in col_map and col_map["Qty"] < len(row_vals) else ""
    
    if part_no and part_no != "-":
        items.append({"partNo": part_no, "description": desc, "qty": qty})

print(f"KO Number: {ko_number}")
print(f"Found {len(items)} valid items.")
print("First 3 items:")
print(items[:3])
